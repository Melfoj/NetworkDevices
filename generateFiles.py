from netmiko import ConnectHandler
from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil
import os
import prFunctions as pr


def boldExcel(worksheet, row, col, text, size = 11): # Helper function for excel
    cell=worksheet.cell(row=row, column=col)
    cell.value=text
    cell.font=Font(bold=True, size=size)


def generateAll(deviceNumber, cityName, address, provider1, provider2): # Generates folder, ARP table and Excel
    # with pattern for reconnecting cables from current production equipment to new devices
    netConnect1 = ConnectHandler(device_type="cisco_ios",
                                host="1.1.1." + str(int(deviceNumber)),# Fill in with your IP, username, password and secret
                                username="cisco",
                                password="cisco",
                                secret="cisco")
    netConnect1.enable()
    path= "Output\\Exp_"+ str(deviceNumber)
    if not os.path.exists(path):
        os.mkdir(path)

    # Create ARP file
    outputArp = netConnect1.send_command("sh arp")
    f = open("Output\\Exp_" + str(deviceNumber) + "\\ARP+" + str(deviceNumber) + ".txt", "a")
    f.write(outputArp)
    f.close()

    # Get port number
    netConnect2 = ConnectHandler(device_type="cisco_ios",
                                host="1.1.1." + str(int(deviceNumber)),
                                username="cisco",
                                password="cisco",
                                secret="cisco")
    netConnect2.enable()
    outputInvSW = netConnect2.send_command("sh inv")
    deviceModel = pr.findWordAfterSpecword(outputInvSW, "PID:")
    if deviceModel is not None:
        deviceModelCorr = deviceModel.replace("+", "-")
        modelSp = deviceModelCorr.split("-")
    if modelSp[0] == "WS" or modelSp[0] - "SM":
        modelSp.pop(0)
    if modelSp[0] == "X":
        modelSp.pop(0)
    portNum = modelSp[1]
    portNum = int("".join(c for c in portNum if c.isdecimal()))

    # Create cable table
    outputMacTable = netConnect2.send_command("sh mac address-table")
    arpArr = outputArp.split(" ")
    i = 0
    while i < len(arpArr):
        if arpArr[i] == "":
            del arpArr[i]
    else:
        i += 1
    i=8
    arpArrS=[]
    while i < len(arpArr):
        arpArrS.append(arpArr[i].rstrip("\n"))
        i+=2
        arpArrS.append(arpArr[i].rstrip("\n"))
        i+=3
    macArr = outputMacTable.split(" ")
    i=0
    while i < len(macArr):
        if macArr[i] == "":
            del macArr[i]
        else:
            i+=1
    i = 11
    macArrS = []
    while i < (len(macArr) - 6):
        macArrS.append(macArr[i].rstrip("\n"))
        i += 2
    temp=macArrS[-1]
    macArrS[-1] = temp.rstrip("\nTotal")
    if portNum == 24:
        shutil.copy("Docs\\PrepovezivanjeTabela24.xlsx",
                    "Output\\Exp_"+str(deviceNumber)+"\\PrepovezivanjeTabela_" + str(deviceNumber) + ".xlsx")
    else:
        shutil.copy("Docs\\PrepovezivanjeTabela.xlsx",
                    "Output\\Exp_"+str(deviceNumber)+"\\PrepovezivanjeTabela_" + str(deviceNumber) + ".xlsx")
    workbook= load_workbook("Output\\Exp_"+str(deviceNumber)+"\\PrepovezivanjeTabela_"+ str(deviceNumber) + ".xlsx")
    worksheet = workbook["Sheet1"]
    boldExcel(worksheet, 1, 1, cityName+", "+address, 20)
    boldExcel(worksheet, 7, 3, provider1)
    boldExcel(worksheet, 8, 3, provider2)
    workbook.save("Output\\Exp_"+str(deviceNumber)+"\\PrepovezivanjeTabela_"+ str(deviceNumber) +".xlsx")
    print(str(int(deviceNumber))+": Files generated.")
